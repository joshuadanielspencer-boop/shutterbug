from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_WORKBOOK = Path("Protestant Chapel Schedule.xlsx")
OUTPUT_JSON = Path("codex_temp/song_frequency_data.json")

# Titles that contain commas and should not be split on those internal commas.
PROTECTED_TITLES = [
    "10,000 Reasons (Bless The Lord)",
    "10,000 Reasons",
    "O Come, O Come, Emmanuel",
    "Come, Thou Long-Expected Jesus",
    "Holy, Holy, Holy!",
    "Breathe On Me, Breath Of God",
    "Lord, Guard and Guide the Men Who Fly",
    "O Sacred Head, Now Wounded",
]

# One cell is missing a delimiter between two titles, so split it manually.
MANUAL_SPLITS = {
    ("C (24-25)", 3): [
        "Come Thou Long Expected Jesus",
        "Come Behold the Wondrous Mystery",
        "O Come O Come Emmanuel",
    ],
}

# Normalize obvious title variants so counts roll up meaningfully.
CANONICAL_TITLES = {
    "O Come, O Come, Emmanuel": "O Come O Come Emmanuel",
    "Come, Thou Long-Expected Jesus": "Come Thou Long Expected Jesus",
    "Holy, Holy, Holy!": "Holy Holy Holy",
    "10,000 Reasons (Bless The Lord)": "10,000 Reasons",
    "O Praise The Name": "O Praise the Name",
    "O Praise the Name (Anastasis)": "O Praise the Name",
    "Come Christians Join to Sing": "Come Christian Join to Sing",
    "Jesus Paid it All": "Jesus Paid It All",
    "Abide With Me": "Abide with Me",
    "His Mercy is More": "His Mercy Is More",
    "All Creatures Of Our God And King": "All Creatures of Our God and King",
    "How Deep The Father's Love For Us": "How Deep the Father's Love for Us",
    "Before The Throne Of God Above": "Before the Throne of God Above",
    "Joy To The World": "Joy to the World",
    "It Came Upon The Midnight Clear": "It Came Upon a Midnight Clear",
    "Hark The Herald Angels Sing": "Hark the Herald Angels Sing",
    "Crown Him With Many Crowns": "Crown Him with Many Crowns",
    "Come Thou Fount Of Every Blessing": "Come Thou Fount",
    "Take My Life": "Take My Life and Let it Be",
    "At The Cross (Love Ran Red)": "At the Cross (Love Ran Red)",
    "When I Survey The Wondrous Cross": "When I Survey the Wondrous Cross",
    "Not to Us": "Not To Us",
}


def split_songs(raw_value: str, key: tuple[str, int] | None = None) -> list[str]:
    if key in MANUAL_SPLITS:
        return MANUAL_SPLITS[key]

    text = str(raw_value).strip()
    placeholders: dict[str, str] = {}

    for index, title in enumerate(PROTECTED_TITLES):
        placeholder = f"__PROTECTED_{index}__"
        if title in text:
            text = text.replace(title, placeholder)
            placeholders[placeholder] = title

    parts = [part.strip() for part in text.split(",") if part.strip()]
    restored: list[str] = []
    for part in parts:
        for placeholder, title in placeholders.items():
            part = part.replace(placeholder, title)
        restored.append(CANONICAL_TITLES.get(part.strip(), part.strip()))
    return restored


def normalize_key(title: str) -> str:
    compact = re.sub(r"\s+", " ", title.replace("’", "'")).strip()
    return compact.lower()


def main() -> None:
    workbook = load_workbook(SOURCE_WORKBOOK, data_only=True)

    counts: Counter[str] = Counter()
    display_names: dict[str, str] = {}
    by_sheet: defaultdict[str, set[str]] = defaultdict(set)

    for sheet in workbook.worksheets:
        for row in range(2, sheet.max_row + 1):
            raw_value = sheet[f"I{row}"].value
            if not raw_value:
                continue

            for song in split_songs(raw_value, (sheet.title, row)):
                normalized = normalize_key(song)
                counts[normalized] += 1
                display_names.setdefault(normalized, song)
                by_sheet[normalized].add(sheet.title)

    rows = [
        {
            "song_title": display_names[key],
            "count": count,
            "tabs_present": ", ".join(sorted(by_sheet[key])),
        }
        for key, count in counts.items()
    ]
    rows.sort(key=lambda item: (-item["count"], item["song_title"].lower()))

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(rows, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
